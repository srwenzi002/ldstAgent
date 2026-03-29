from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from slack_excel_bot.template_loader import TemplateLoaderError, load_mapping, load_registry_entry
from slack_excel_bot.template_schema import ItemFieldMapping, MissingBehavior, TemplateMapping


class ExcelWriteError(Exception):
    def __init__(self, code: str, message: str, details: dict | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass
class DraftWriteResult:
    draft_file_id: str
    output_path: str
    template_id: str


class ExcelWriter:
    def __init__(self, package_dir: Path, draft_dir: Path):
        self.package_dir = package_dir
        self.registry_path = package_dir / "templates" / "registry.json"
        self.draft_dir = draft_dir
        self.draft_dir.mkdir(parents=True, exist_ok=True)
        self._jst = timezone(timedelta(hours=9))

    def write_draft(self, template_id: str, canonical_model: dict[str, Any]) -> DraftWriteResult:
        try:
            entry = load_registry_entry(template_id=template_id, registry_path=self.registry_path)
        except TemplateLoaderError as exc:
            raise ExcelWriteError(exc.code, exc.message, exc.details) from exc

        if not entry.enabled:
            raise ExcelWriteError("TEMPLATE_DISABLED", "Template is disabled.", {"template_id": template_id})

        template_file_path = self._resolve_existing_path(self.package_dir / entry.file_path)
        if template_file_path is None:
            raise ExcelWriteError(
                "TEMPLATE_FILE_NOT_FOUND",
                "Template workbook file does not exist.",
                {"template_id": template_id, "file_path": entry.file_path},
            )

        mapping_path = self.package_dir / entry.mapping_path
        try:
            mapping = load_mapping(mapping_path)
        except TemplateLoaderError as exc:
            raise ExcelWriteError(exc.code, exc.message, exc.details) from exc

        workbook = load_workbook(template_file_path)
        if mapping.sheet not in workbook.sheetnames:
            raise ExcelWriteError("SHEET_NOT_FOUND", "Mapped sheet is not present in workbook.", {"sheet": mapping.sheet})
        sheet = workbook[mapping.sheet]

        self._apply_constants(sheet, mapping)
        self._apply_single_fields(sheet, mapping, canonical_model)
        self._apply_items(sheet, mapping, canonical_model, template_id)

        draft_file_id = self._build_draft_file_id(template_id, template_file_path, canonical_model)
        output_path = self.draft_dir / f"{draft_file_id}.xlsx"
        workbook.save(output_path)

        return DraftWriteResult(
            draft_file_id=draft_file_id,
            output_path=str(output_path),
            template_id=template_id,
        )

    def _apply_constants(self, sheet, mapping: TemplateMapping) -> None:
        for constant in mapping.constants:
            sheet[constant.cell] = constant.value

    def _apply_single_fields(self, sheet, mapping: TemplateMapping, canonical: dict[str, Any]) -> None:
        for field_map in mapping.single_fields:
            value, exists = self._resolve_path(canonical, field_map.field_path)
            if exists:
                sheet[field_map.cell] = self._coerce_value_by_type(value, field_map.value_type)
                continue

            behavior = self._resolve_missing_behavior(mapping, field_map.field_path, field_map.missing)
            if behavior == MissingBehavior.BLANK:
                sheet[field_map.cell] = None
                continue

            raise ExcelWriteError(
                "FIELD_MISSING",
                "Required field is missing for template mapping.",
                {"field_path": field_map.field_path, "cell": field_map.cell},
            )

    def _apply_items(
        self,
        sheet,
        mapping: TemplateMapping,
        canonical: dict[str, Any],
        template_id: str,
    ) -> None:
        if mapping.items is None:
            return

        items_value, exists = self._resolve_path(canonical, mapping.items.path)
        if not exists:
            if self._resolve_missing_behavior(mapping, mapping.items.path, mapping.items.missing) == MissingBehavior.BLANK:
                return
            raise ExcelWriteError("FIELD_MISSING", "Required items field is missing.", {"field_path": mapping.items.path})

        if not isinstance(items_value, list):
            raise ExcelWriteError(
                "INVALID_ITEMS_TYPE",
                "Mapped items path must point to a list.",
                {"field_path": mapping.items.path, "actual_type": type(items_value).__name__},
            )

        if len(items_value) > mapping.items.max_rows:
            raise ExcelWriteError(
                "ITEMS_ROW_LIMIT_EXCEEDED",
                "Item rows exceed template max rows.",
                {
                    "template_id": template_id,
                    "max_rows": mapping.items.max_rows,
                    "actual_rows": len(items_value),
                },
            )

        field_specs = self._build_item_field_specs(mapping.items.columns, mapping.items.fields)
        used_row_indexes: set[int] = set()
        for index, item in enumerate(items_value):
            if not isinstance(item, dict):
                raise ExcelWriteError(
                    "INVALID_ITEM_ROW",
                    "Each item row must be an object.",
                    {"index": index, "actual_type": type(item).__name__},
                )
            base_row = self._resolve_base_row_for_item(mapping, item, index, used_row_indexes)
            for item_field, field_spec in field_specs.items():
                value, field_exists = self._resolve_path(item, item_field)
                if not field_exists:
                    behavior = self._resolve_missing_behavior(
                        mapping,
                        f"{mapping.items.path}.{item_field}",
                        field_spec.missing or mapping.items.missing,
                    )
                    if behavior == MissingBehavior.ERROR:
                        raise ExcelWriteError(
                            "FIELD_MISSING",
                            "Required item field is missing.",
                            {"field_path": f"{mapping.items.path}.{item_field}", "row_index": index},
                        )
                    value = None
                if isinstance(value, bool) and (field_spec.true_value is not None or field_spec.false_value is not None):
                    value = field_spec.true_value if value else field_spec.false_value
                sheet[f"{field_spec.column}{base_row + field_spec.row_offset}"] = self._coerce_value_by_type(
                    value, field_spec.value_type
                )

    def _resolve_base_row_for_item(
        self,
        mapping: TemplateMapping,
        item: dict[str, Any],
        item_index: int,
        used_row_indexes: set[int],
    ) -> int:
        assert mapping.items is not None
        if mapping.items.row_index_mode == "sequence":
            return mapping.items.start_row + (item_index * mapping.items.row_stride)

        value, exists = self._resolve_path(item, mapping.items.row_index_field)
        if not exists:
            raise ExcelWriteError(
                "FIELD_MISSING",
                "Required item row index field is missing.",
                {"field_path": f"{mapping.items.path}.{mapping.items.row_index_field}", "row_index": item_index},
            )

        row_index = self._parse_day_of_month(value)
        if row_index is None:
            raise ExcelWriteError(
                "INVALID_ITEM_ROW_INDEX",
                "Item row index value is invalid.",
                {"value": value, "row_index": item_index},
            )

        if row_index < 1 or row_index > mapping.items.max_rows:
            raise ExcelWriteError(
                "ITEM_ROW_INDEX_OUT_OF_RANGE",
                "Item row index is out of supported range for template.",
                {"value": value, "min_index": 1, "max_index": mapping.items.max_rows},
            )

        if row_index in used_row_indexes:
            raise ExcelWriteError(
                "DUPLICATE_ITEM_ROW_INDEX",
                "Duplicate row index detected in items.",
                {"value": value, "row_index": item_index},
            )
        used_row_indexes.add(row_index)
        return mapping.items.start_row + ((row_index - 1) * mapping.items.row_stride)

    def _build_draft_file_id(self, template_id: str, template_file_path: Path, canonical_model: dict[str, Any]) -> str:
        base_name = template_file_path.stem or template_id
        yyyymm = self._resolve_yyyymm(canonical_model)
        name = self._resolve_value(canonical_model, ("employee.name",)) or "氏名未入力"
        employee_id = self._resolve_value(canonical_model, ("employee.employee_id",)) or "社員番号未入力"
        department_code = self._resolve_value(canonical_model, ("employee.department_code", "employee.department")) or "部署コード未入力"

        replaced = base_name
        for token, value in {
            "YYYYMM": yyyymm,
            "氏名": name,
            "社員番号": employee_id,
            "部署コード": department_code,
            "部署コード": department_code,
        }.items():
            replaced = replaced.replace(token, str(value))

        safe = self._sanitize_filename(replaced) or template_id
        timestamp = datetime.now(self._jst).strftime("%Y%m%d%H%M%S")
        return f"{safe}_{timestamp}_{uuid4().hex[:6]}"

    def _resolve_yyyymm(self, canonical_model: dict[str, Any]) -> str:
        year = self._to_int(self._resolve_value(canonical_model, ("year",)))
        month = self._to_int(self._resolve_value(canonical_model, ("month",)))
        if year is not None and month is not None and 1 <= month <= 12:
            return f"{year:04d}{month:02d}"
        return datetime.now(self._jst).strftime("%Y%m")

    @staticmethod
    def _resolve_path(payload: dict[str, Any], path: str) -> tuple[Any, bool]:
        current: Any = payload
        for part in path.split("."):
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                return None, False
        return current, True

    @staticmethod
    def _resolve_value(payload: dict[str, Any], candidate_paths: tuple[str, ...]) -> str | None:
        for path in candidate_paths:
            value, exists = ExcelWriter._resolve_path(payload, path)
            if exists and value is not None and str(value).strip():
                return str(value).strip()
        return None

    @staticmethod
    def _resolve_missing_behavior(
        mapping: TemplateMapping,
        field_path: str,
        override: MissingBehavior | None,
    ) -> MissingBehavior:
        if override is not None:
            return override
        if field_path in mapping.missing_strategy.fields:
            return mapping.missing_strategy.fields[field_path]
        return mapping.missing_strategy.default

    @staticmethod
    def _build_item_field_specs(columns: dict[str, str], fields: dict[str, ItemFieldMapping]) -> dict[str, ItemFieldMapping]:
        specs = {field_path: ItemFieldMapping(column=column) for field_path, column in columns.items()}
        specs.update(fields)
        return specs

    @staticmethod
    def _to_int(value: Any) -> int | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, str) and value.strip().isdigit():
            return int(value.strip())
        return None

    @staticmethod
    def _parse_day_of_month(value: Any) -> int | None:
        if isinstance(value, int):
            return value
        if isinstance(value, date):
            return value.day
        if isinstance(value, str):
            stripped = value.strip()
            if stripped.isdigit():
                return int(stripped)
            try:
                return datetime.fromisoformat(stripped.replace("/", "-")).day
            except ValueError:
                return None
        return None

    @staticmethod
    def _coerce_value_by_type(value: Any, value_type: str) -> Any:
        if value is None:
            return None
        if value_type == "auto":
            return ExcelWriter._coerce_auto_value(value)
        if value_type == "string":
            return str(value)
        if value_type == "int":
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return int(value)
            if isinstance(value, str) and value.strip():
                try:
                    return int(value.strip())
                except ValueError:
                    return value
            return value
        if value_type == "float":
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return float(value)
            if isinstance(value, str) and value.strip():
                try:
                    return float(value.strip())
                except ValueError:
                    return value
            return value
        if value_type == "date":
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
                    try:
                        return datetime.strptime(value.strip(), fmt).date()
                    except ValueError:
                        pass
            return value
        if value_type == "time":
            if isinstance(value, datetime):
                return value.time()
            if isinstance(value, time):
                return value
            if isinstance(value, str):
                for fmt in ("%H:%M", "%H:%M:%S"):
                    try:
                        return datetime.strptime(value.strip(), fmt).time()
                    except ValueError:
                        pass
            return value
        if value_type == "datetime":
            if isinstance(value, datetime):
                return value
            if isinstance(value, str) and value.strip():
                try:
                    return datetime.fromisoformat(value.strip().replace("/", "-"))
                except ValueError:
                    return value
        return value

    @staticmethod
    def _coerce_auto_value(value: Any) -> Any:
        if isinstance(value, (list, tuple, set)):
            return "、".join(str(item) for item in value if item is not None)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return value

    @staticmethod
    def _sanitize_filename(raw: str) -> str:
        text = re.sub(r'[\\/:*?"<>|]+', "_", raw)
        text = re.sub(r"\s+", "_", text).strip("._ ")
        return re.sub(r"_+", "_", text)

    @staticmethod
    def _build_path_variants(path: Path) -> list[str]:
        base = str(path)
        variants = [base]
        for form in ("NFC", "NFD"):
            normalized = unicodedata.normalize(form, base)
            if normalized not in variants:
                variants.append(normalized)
        return variants

    @classmethod
    def _resolve_existing_path(cls, path: Path) -> Path | None:
        for variant in cls._build_path_variants(path):
            candidate = Path(variant)
            if candidate.exists():
                return candidate
        return None
