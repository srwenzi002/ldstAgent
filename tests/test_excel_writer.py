from pathlib import Path

from openpyxl import load_workbook

from slack_excel_bot.excel_tools import ExcelToolService
from slack_excel_bot.config import Settings
from slack_excel_bot.ekispert_client import EkispertError, EkispertMcpClient
from slack_excel_bot.excel_writer import ExcelWriteError, ExcelWriter


ROOT = Path(__file__).resolve().parents[1]
PACKAGE_DIR = ROOT / "src" / "slack_excel_bot"


def build_settings(tmp_path: Path) -> Settings:
    return Settings(
        slack_bot_token="xoxb-test",
        slack_app_token="xapp-test",
        openai_api_key="sk-test",
        openai_model="gpt-4.1-mini",
        ekispert_api_token="ek-test",
        port=3000,
        storage_dir=tmp_path,
        default_employee_name="山田太郎",
        default_employee_id="0001",
        default_department="開発本部",
        default_department_code="50",
        default_work_grade=1,
        default_clock_in="09:00",
        default_clock_out="18:00",
        max_concurrent_requests=50,
    )


def test_transport_writer_generates_workbook(tmp_path: Path) -> None:
    writer = ExcelWriter(package_dir=PACKAGE_DIR, draft_dir=tmp_path)
    payload = {
        "employee": {
            "department": "開発本部",
            "employee_id": "0001",
            "name": "山田太郎",
        },
        "items": [
            {
                "travel_date": "2026-03-10",
                "purpose": "客先作業",
                "visit_place": "渋谷オフィス",
                "transport_mode": "電車・バス",
                "route_from": "新宿",
                "route_to": "渋谷",
                "route_line": "JR山手線",
                "one_way_amount": 178,
                "is_round_trip": True,
                "receipt_no": "R-001",
            }
        ],
    }

    result = writer.write_draft("transport_jp_leadingsoft_v1", payload)
    wb = load_workbook(result.output_path)
    ws = wb["精算書（交通費）"]

    assert ws["J4"].value == "開発本部"
    assert ws["N4"].value == "0001"
    assert ws["S4"].value == "山田太郎"
    assert ws["A9"].value is not None
    assert ws["AD9"].value == "●"


def test_transport_tool_generates_workbook(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.generate_transport_sheet(
        {
            "employee": {
                "department": "ソリューション開発部",
                "department_code": "51",
                "employee_id": "0001",
                "name": "山田太郎",
            },
            "items": [
                {
                    "travel_date": "2026-03-10",
                    "purpose": "客先作業",
                    "visit_place": "渋谷オフィス",
                    "transport_mode": "電車・バス",
                    "route_from": "新宿",
                    "route_to": "渋谷",
                    "route_line": "JR山手線",
                    "one_way_amount": 178,
                    "is_round_trip": True,
                    "receipt_no": "R-001",
                }
            ],
        }
    )
    wb = load_workbook(result["output_path"])
    ws = wb["精算書（交通費）"]

    assert ws["J4"].value == "ソリューション開発部"
    assert ws["N4"].value == "0001"
    assert ws["S4"].value == "山田太郎"
    assert ws["F9"].value == "客先作業"
    assert ws["K9"].value == "電車・バス"
    assert ws["AA9"].value == 178
    assert ws["AD9"].value == "●"
    assert result["title"] == "精算書_集（交通費）_山田太郎_202603"


def test_transport_tool_applies_default_purpose_and_round_trip(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.generate_transport_sheet(
        {
            "employee": {
                "department": "ソリューション開発部",
                "department_code": "51",
                "employee_id": "0001",
                "name": "山田太郎",
            },
            "items": [
                {
                    "travel_date": "2026-03-29",
                    "purpose": None,
                    "visit_place": None,
                    "transport_mode": "電車・バス",
                    "route_from": "青砥",
                    "route_to": "青物横丁",
                    "route_line": None,
                    "one_way_amount": 651,
                    "is_round_trip": None,
                    "receipt_no": None,
                }
            ],
        }
    )
    wb = load_workbook(result["output_path"])
    ws = wb["精算書（交通費）"]

    assert ws["F9"].value == "営業活動"
    assert ws["AD9"].value in (None, "")


def test_writer_raises_for_missing_required_field(tmp_path: Path) -> None:
    writer = ExcelWriter(package_dir=PACKAGE_DIR, draft_dir=tmp_path)
    payload = {
        "employee": {
            "department": "開発本部",
            "employee_id": "0001",
        },
        "items": [],
    }

    try:
        writer.write_draft("transport_jp_leadingsoft_v1", payload)
    except ExcelWriteError as exc:
        assert exc.code == "FIELD_MISSING"
        assert exc.details["field_path"] == "employee.name"
    else:
        raise AssertionError("Expected ExcelWriteError")


def test_attendance_tool_expands_weekdays(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.generate_attendance_sheet(
        {
            "year": 2026,
            "month": 3,
            "employee": {"department_code": "51", "employee_id": "0001", "name": "山田太郎"},
            "days": [
                {"day": 2, "work_grade": 2, "clock_in": "09:00", "clock_out": "17:30"},
                {"day": 3, "work_grade": 2, "clock_in": "09:00", "clock_out": "17:30"},
            ],
        }
    )
    wb = load_workbook(result["output_path"])
    ws = wb["勤務状況表"]

    assert ws["AE4"].value == "0001"
    assert ws["AF4"].value == "山田太郎"
    assert ws["F12"].value is None
    assert ws["G12"].value is None
    assert ws["F13"].value is not None
    assert ws["G13"].value is not None
    assert ws["E13"].value == 2
    assert result["title"] == "Ldjpw668_2603_51_0001_山田太郎"


def test_attendance_tool_writes_half_day_leave_directly(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.generate_attendance_sheet(
        {
            "year": 2026,
            "month": 2,
            "employee": {"department_code": "51", "employee_id": "0001", "name": "山田太郎"},
            "days": [
                {
                    "day": 6,
                    "leave_item_no": 2,
                    "work_grade": 2,
                    "clock_in": "13:00",
                    "clock_out": "18:00",
                }
            ],
        }
    )
    wb = load_workbook(result["output_path"])
    ws = wb["勤務状況表"]

    assert ws["K17"].value == 2
    assert ws["E17"].value == 2
    assert ws["F17"].value is not None
    assert ws["G17"].value is not None


def test_personal_expense_tool_generates_workbook(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.generate_personal_expense_sheet(
        {
            "employee": {
                "department": "ソリューション開発部",
                "department_code": "51",
                "employee_id": "0001",
                "name": "山田太郎",
            },
            "items": [
                {
                    "expense_date": "2026-03-15",
                    "purpose": "会議費",
                    "amount_jpy": 5500,
                    "payee_name": "渋谷カフェ",
                    "description": "顧客打合せの飲食代",
                    "burden_department": "ソリューション開発部",
                    "project_code_name": "SD0001：SD部門経費",
                    "counterparty_company": "株式会社サンプル",
                    "counterparty_attendees": "佐藤様",
                    "counterparty_count": 1,
                    "internal_attendees": "山田太郎",
                    "internal_count": 1,
                }
            ],
        }
    )
    wb = load_workbook(result["output_path"])
    ws = wb["経費（個人）精算書"]

    assert ws["J5"].value == "ソリューション開発部"
    assert ws["N5"].value == "0001"
    assert ws["R5"].value == "山田太郎"
    assert ws["J10"].value == "会議費"
    assert ws["Y10"].value == 5500
    assert ws["A12"].value == "渋谷カフェ"
    assert ws["I12"].value == "顧客打合せの飲食代"
    assert result["title"] == "精算書_集（個人経費立替）_山田太郎_202603"


def test_ekispert_route_parser_extracts_candidates() -> None:
    payload = {
        "ResultSet": {
            "Course": [
                {
                    "Teiki": {"DisplayRoute": "青物横丁--京急本線--品川--ＪＲ山手線内回り--浜松町"},
                    "Price": [
                        {"kind": "FareSummary", "Oneway": "350", "Round": "700"},
                    ],
                    "Route": {
                        "timeOnBoard": "10",
                        "timeOther": "12",
                        "timeWalk": "0",
                        "transferCount": "1",
                        "Point": [
                            {"Station": {"Name": "青物横丁"}},
                            {"Station": {"Name": "品川"}},
                            {"Station": {"Name": "浜松町"}},
                        ],
                    },
                }
            ]
        }
    }

    options = EkispertMcpClient._parse_route_options(payload, top_k=3)

    assert len(options) == 1
    assert options[0].route_line == "青物横丁 -> 京急本線 -> 品川 -> ＪＲ山手線内回り -> 浜松町"
    assert options[0].route_summary == "青物横丁 -> 品川 -> 浜松町"
    assert options[0].one_way_amount == 350
    assert options[0].total_minutes == 22
    assert options[0].transfer_count == 1


def test_transport_route_lookup_returns_structured_options(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    class StubClient:
        def search_route_options(self, *, route_from: str, route_to: str, top_k: int, travel_date: str | None):
            assert route_from == "青物横丁"
            assert route_to == "浜松町"
            assert top_k == 2
            assert travel_date == "2026-03-29"
            return [
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "1",
                            "route_summary": "青物横丁 -> 品川 -> 浜松町",
                            "route_line": "青物横丁 -> 京急本線 -> 品川 -> ＪＲ山手線内回り -> 浜松町",
                            "one_way_amount": 350,
                            "total_minutes": 22,
                            "transfer_count": 1,
                        }
                    },
                )()
            ]

    service.ekispert_client = StubClient()

    result = service.lookup_transport_route_options(
        {
            "travel_date": "2026-03-29",
            "route_from": "青物横丁",
            "route_to": "浜松町",
            "top_k": 2,
        }
    )

    assert result["ok"] is True
    assert result["title"] == "交通経路候補"
    assert result["options"][0]["one_way_amount"] == 350
    assert result["options"][0]["transfer_count"] == 1


def test_analyze_expense_evidence_transport_fills_missing_fields_and_default_mode(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.analyze_expense_evidence(
        {
            "expense_type": "transport",
            "document_kind": "transport_screenshot",
            "travel_date": "2026-03-29",
            "route_from": "青物横丁",
            "route_to": "浜松町",
            "route_line": None,
            "one_way_amount": None,
            "transport_mode": None,
            "is_round_trip": None,
            "purpose": None,
            "expense_date": None,
            "amount_jpy": None,
            "payee_name": None,
            "description": None,
            "confidence": "medium",
            "evidence_sources": ["text", "image"],
            "missing_fields": [],
            "notes": "截图里没看清金额",
        }
    )

    assert result["title"] == "証憑の解析結果"
    assert result["transport_mode"] == "電車・バス"
    assert result["missing_fields"] == ["one_way_amount", "route_line"]
    assert result["evidence_sources"] == ["text", "image"]


def test_analyze_expense_evidence_personal_expense_uses_personal_fields(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.analyze_expense_evidence(
        {
            "expense_type": "personal_expense",
            "document_kind": "receipt",
            "travel_date": None,
            "route_from": None,
            "route_to": None,
            "route_line": None,
            "one_way_amount": None,
            "transport_mode": None,
            "is_round_trip": None,
            "purpose": None,
            "expense_date": "2026-03-29",
            "amount_jpy": 1200,
            "payee_name": "コンビニA",
            "description": "文房具購入",
            "confidence": "high",
            "evidence_sources": ["image"],
            "missing_fields": ["travel_date"],
            "notes": None,
        }
    )

    assert result["missing_fields"] == []
    assert result["amount_jpy"] == 1200
    assert result["expense_type"] == "personal_expense"
    assert result["confidence"] == "high"


def test_analyze_expense_evidence_transport_items_clear_missing_fields(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    result = service.analyze_expense_evidence(
        {
            "expense_type": "transport",
            "document_kind": "transport_screenshot",
            "travel_date": None,
            "route_from": None,
            "route_to": None,
            "route_line": None,
            "one_way_amount": None,
            "transport_events": [
                {
                    "travel_date": "2026-03-21",
                    "event_kind": "定",
                    "event_role": "pass_entry_or_exit",
                    "station_or_merchant": "青物横丁",
                    "amount_jpy": -510,
                    "balance_jpy": 2553,
                    "paired_group": None,
                    "confidence": "medium",
                    "notes": "定期関連イベントなので通常の片道精算対象外",
                },
                {
                    "travel_date": "2026-03-22",
                    "event_kind": "物販",
                    "event_role": "shopping",
                    "station_or_merchant": None,
                    "amount_jpy": -500,
                    "balance_jpy": 1844,
                    "paired_group": None,
                    "confidence": "medium",
                    "notes": "物販は交通費ではない",
                },
            ],
            "transport_items": [
                {
                    "travel_date": "2026-03-15",
                    "route_from": "押上",
                    "route_to": "青砥",
                    "one_way_amount": 199,
                    "route_line": None,
                    "transport_mode": None,
                    "is_round_trip": None,
                    "purpose": None,
                    "confidence": "medium",
                    "notes": "入/出记录配对推断",
                },
                {
                    "travel_date": "2026-03-15",
                    "route_from": "新高円寺",
                    "route_to": "茅場町",
                    "one_way_amount": 252,
                    "route_line": None,
                    "transport_mode": None,
                    "is_round_trip": None,
                    "purpose": None,
                    "confidence": "medium",
                    "notes": "入/出记录配对推断",
                },
            ],
            "transport_mode": None,
            "is_round_trip": None,
            "purpose": None,
            "expense_date": None,
            "amount_jpy": None,
            "payee_name": None,
            "description": None,
            "confidence": "medium",
            "evidence_sources": ["image"],
            "missing_fields": ["travel_date"],
            "notes": "来自交通卡履历截图",
        }
    )

    assert result["missing_fields"] == []
    assert result["transport_events"][0]["event_kind"] == "定"
    assert result["transport_events"][0]["event_role"] == "pass_entry_or_exit"
    assert result["transport_events"][1]["event_kind"] == "物販"
    assert result["transport_items"][0]["transport_mode"] == "電車・バス"
    assert result["transport_items"][1]["one_way_amount"] == 252


def test_transport_route_batch_lookup_returns_match_and_candidates(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    class StubClient:
        def search_route_options(self, *, route_from: str, route_to: str, top_k: int, travel_date: str | None):
            assert top_k == 2
            return [
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "1",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 路線A -> {route_to}",
                            "one_way_amount": 272,
                            "total_minutes": 18,
                            "transfer_count": 0,
                        }
                    },
                )(),
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "2",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 路線B -> {route_to}",
                            "one_way_amount": 310,
                            "total_minutes": 24,
                            "transfer_count": 1,
                        }
                    },
                )(),
            ]

    service.ekispert_client = StubClient()

    result = service.lookup_transport_route_batch(
        {
            "items": [
                {
                    "travel_date": "2026-03-27",
                    "route_from": "青砥",
                    "route_to": "京成上野",
                    "one_way_amount": 272,
                    "route_line": None,
                }
            ],
            "top_k": 2,
        }
    )

    assert result["title"] == "交通経路の一括確認結果"
    assert result["items"][0]["matched_option"]["one_way_amount"] == 272
    assert result["items"][0]["final_one_way_amount"] == 272
    assert result["items"][0]["match_type"] == "exact"
    assert result["items"][0]["should_prompt_user"] is False
    assert result["resolved_items"][0]["route_line"] == "青砥 -> 路線A -> 京成上野"
    assert len(result["items"][0]["options"]) == 2


def test_transport_route_batch_lookup_tolerates_single_item_failure(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    class StubClient:
        def search_route_options(self, *, route_from: str, route_to: str, top_k: int, travel_date: str | None):
            if route_from == "八景":
                raise EkispertError('{"status": 400, "message": "駅名が見つかりません。(八景)"}')
            return [
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "1",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 路線A -> {route_to}",
                            "one_way_amount": 272,
                            "total_minutes": 18,
                            "transfer_count": 0,
                        }
                    },
                )()
            ]

    service.ekispert_client = StubClient()

    result = service.lookup_transport_route_batch(
        {
            "items": [
                {
                    "travel_date": "2026-03-27",
                    "route_from": "八景",
                    "route_to": "青砥",
                    "one_way_amount": None,
                    "route_line": None,
                },
                {
                    "travel_date": "2026-03-27",
                    "route_from": "青砥",
                    "route_to": "京成上野",
                    "one_way_amount": 272,
                    "route_line": None,
                },
            ],
            "top_k": 2,
        }
    )

    assert result["ok"] is True
    assert result["has_partial_failures"] is True
    assert result["items"][0]["status"] == "query_error"
    assert "八景" in result["items"][0]["error"]
    assert result["items"][1]["status"] == "ok"


def test_transport_route_batch_lookup_uses_image_amount_for_near_ic_fare(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    class StubClient:
        def search_route_options(self, *, route_from: str, route_to: str, top_k: int, travel_date: str | None):
            return [
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "1",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 直通 -> {route_to}",
                            "one_way_amount": 280,
                            "total_minutes": 25,
                            "transfer_count": 0,
                        }
                    },
                )(),
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "2",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 迂回 -> {route_to}",
                            "one_way_amount": 360,
                            "total_minutes": 33,
                            "transfer_count": 1,
                        }
                    },
                )(),
            ]

    service.ekispert_client = StubClient()

    result = service.lookup_transport_route_batch(
        {
            "items": [
                {
                    "travel_date": "2026-03-27",
                    "route_from": "京成上野",
                    "route_to": "青砥",
                    "one_way_amount": 272,
                    "route_line": None,
                }
            ],
            "top_k": 3,
        }
    )

    assert result["items"][0]["match_type"] == "near_ic_fare"
    assert result["items"][0]["matched_option"]["one_way_amount"] == 280
    assert result["items"][0]["final_one_way_amount"] == 272
    assert result["items"][0]["should_prompt_user"] is False
    assert result["resolved_items"][0]["route_line"] == "京成上野 -> 直通 -> 青砥"


def test_transport_route_batch_lookup_prompts_when_multiple_close_candidates_exist(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    class StubClient:
        def search_route_options(self, *, route_from: str, route_to: str, top_k: int, travel_date: str | None):
            return [
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "1",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 路線A -> {route_to}",
                            "one_way_amount": 210,
                            "total_minutes": 21,
                            "transfer_count": 0,
                        }
                    },
                )(),
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "2",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 路線B -> {route_to}",
                            "one_way_amount": 209,
                            "total_minutes": 22,
                            "transfer_count": 0,
                        }
                    },
                )(),
            ]

    service.ekispert_client = StubClient()

    result = service.lookup_transport_route_batch(
        {
            "items": [
                {
                    "travel_date": "2026-03-22",
                    "route_from": "浅草橋",
                    "route_to": "千駄ケ谷",
                    "one_way_amount": 209,
                    "route_line": None,
                }
            ],
            "top_k": 3,
        }
    )

    assert result["items"][0]["matched_option"] is None
    assert result["items"][0]["recommended_option"]["one_way_amount"] == 209
    assert result["items"][0]["should_prompt_user"] is True
    assert result["items"][0]["prompt_reason"] == "multiple_close_candidates"
    assert len(result["resolved_items"]) == 0
    assert len(result["needs_confirmation"]) == 1


def test_transport_route_batch_lookup_merges_inverse_same_day_items_as_round_trip(tmp_path: Path) -> None:
    settings = build_settings(tmp_path)
    service = ExcelToolService(settings)

    class StubClient:
        def search_route_options(self, *, route_from: str, route_to: str, top_k: int, travel_date: str | None):
            return [
                type(
                    "StubOption",
                    (),
                    {
                        "as_dict": lambda self: {
                            "option_id": "1",
                            "route_summary": f"{route_from} -> {route_to}",
                            "route_line": f"{route_from} -> 京成本線 -> {route_to}",
                            "one_way_amount": 280,
                            "total_minutes": 25,
                            "transfer_count": 0,
                        }
                    },
                )()
            ]

    service.ekispert_client = StubClient()

    result = service.lookup_transport_route_batch(
        {
            "items": [
                {
                    "travel_date": "2026-03-27",
                    "route_from": "京成上野",
                    "route_to": "青砥",
                    "one_way_amount": 272,
                    "route_line": None,
                },
                {
                    "travel_date": "2026-03-27",
                    "route_from": "青砥",
                    "route_to": "京成上野",
                    "one_way_amount": 272,
                    "route_line": None,
                },
            ],
            "top_k": 2,
        }
    )

    assert len(result["resolved_items"]) == 1
    assert result["resolved_items"][0]["route_from"] == "京成上野"
    assert result["resolved_items"][0]["route_to"] == "青砥"
    assert result["resolved_items"][0]["is_round_trip"] is True
    assert len(result["round_trip_suggestions"]) == 1
    assert result["round_trip_suggestions"][0]["merged_item_ids"] == ["1", "2"]
